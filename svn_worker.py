import os
import re
import subprocess
from datetime import datetime
from PySide6.QtCore import QThread, Signal
import openpyxl

class SVNWorker(QThread):
    progress = Signal(int)
    finished = Signal(bool, str)
    log_message = Signal(str)
    
    def __init__(self, url, username, password, excel_path, filter_patterns=None):
        super().__init__()
        self.url = url.rstrip('/')  # Remove trailing slash from URL
        self.username = username
        self.password = password
        self.excel_path = excel_path
        self.filter_patterns = filter_patterns or []
        self.translations = {}  # Will be set by set_translations
        
    def set_translations(self, translations):
        """Set the translations dictionary for the current language"""
        self.translations = translations
        
    def tr(self, key, **kwargs):
        """Translate the given key with format arguments"""
        text = self.translations.get(key, key)
        return text.format(**kwargs) if kwargs else text
        
    def log(self, message):
        self.log_message.emit(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
        
    def is_file_matched(self, file_path):
        if not self.filter_patterns:
            return True
        
        for pattern in self.filter_patterns:
            # Convert wildcard pattern to regex
            regex_pattern = pattern.replace('.', '\\.').replace('*', '.*')
            if re.search(regex_pattern, file_path, re.IGNORECASE):
                return True
        return False
        
    def parse_svn_line(self, line):
        """Parse a single line of SVN output"""
        try:
            pattern = r'^\s*(\d+)\s+(\S+)\s+(?:(\d+)\s+)?([^\s]+ \d+(?:\s+\d+|\s+\d+:\d+))\s+(.+)$'
            match = re.match(pattern, line.strip())
            
            if match:
                groups = match.groups()
                revision = groups[0]
                author = groups[1]
                size = groups[2] if groups[2] else ''
                date_time = groups[3]
                path = groups[4].strip()
                
                self.log(self.tr('log_parse_result',
                    revision=revision,
                    author=author,
                    size=size,
                    date_time=date_time,
                    path=path
                ))
                
                return {
                    'revision': revision,
                    'author': author,
                    'size': size,
                    'date': date_time,
                    'time': '',  # Date time is already merged in date field
                    'path': path
                }
        except Exception as e:
            self.log(self.tr('log_parse_failed', line=line, error=str(e)))
        return None
        
    def run_svn_command(self, command):
        try:
            safe_command = command.replace(self.password, '*' * len(self.password))
            self.log(self.tr('log_executing_command', command=safe_command))
            
            # Set environment variables for handling Chinese characters
            env = os.environ.copy()
            env['LANG'] = 'zh_CN.UTF-8'
            env['LC_ALL'] = 'zh_CN.UTF-8'
            
            # Try to execute command directly with cmd
            if 'list' in command:
                command = f'cmd /c {command}'
            
            process = subprocess.Popen(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding='gb2312',
                shell=True,
                env=env
            )
            stdout, stderr = process.communicate()
            
            if process.returncode != 0:
                error_msg = stderr if stderr else "Unknown error"
                self.log(self.tr('log_command_failed', code=process.returncode))
                self.log(self.tr('log_error_message', error=error_msg))
                raise Exception(self.tr('error_svn', error=error_msg))
            
            if not stdout and not stderr:
                self.log(self.tr('log_command_no_output'))
                raise Exception(self.tr('error_no_output'))
            
            self.log(self.tr('log_command_success'))
            return stdout
        except UnicodeDecodeError:
            self.log(self.tr('log_encoding_error'))
            try:
                process = subprocess.Popen(
                    command,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    encoding='gbk',
                    shell=True,
                    env=env
                )
                stdout, stderr = process.communicate()
                if stdout:
                    return stdout
            except Exception as e:
                self.log(self.tr('log_encoding_retry_failed', error=str(e)))
            raise Exception(self.tr('error_encoding'))
        except Exception as e:
            self.log(self.tr('log_error_occurred', error=str(e)))
            raise
            
    def run(self):
        try:
            self.log(self.tr('log_starting_export'))
            
            entries = []
            try:
                if not self.url.startswith(('http://', 'https://', 'svn://', 'file:///')):
                    raise Exception(self.tr('error_invalid_url'))

                command = f'svn list "{self.url}" --username "{self.username}" --password "{self.password}" -R --verbose'
                
                output = self.run_svn_command(command)
                self.log(self.tr('log_file_list_success'))
                
                if not output:
                    raise Exception(self.tr('error_no_output_check'))
                
                lines = output.strip().split('\n')
                
                for line in lines:
                    if not line or not line.strip():
                        continue
                    
                    # Parse SVN output line
                    parsed = self.parse_svn_line(line)
                    if not parsed:
                        continue
                        
                    path = parsed['path']
                    if path.endswith('/'):
                        continue
                        
                    # Check if file matches filter
                    if not self.is_file_matched(path):
                        continue
                        
                    # Get filename and directory path
                    file_name = os.path.basename(path)
                    dir_path = os.path.dirname(path)
                    
                    # Build complete SVN path
                    if dir_path:
                        full_svn_path = f"{self.url}/{dir_path}"
                    else:
                        full_svn_path = self.url
                    
                    # Only handle duplicate slashes in path part, protect protocol part
                    if '://' in full_svn_path:
                        protocol, path = full_svn_path.split('://', 1)
                        full_svn_path = f"{protocol}://{path.replace('//', '/')}"
                    else:
                        full_svn_path = full_svn_path.replace('//', '/')
                    
                    entries.append({
                        'file_name': file_name,
                        'dir_path': full_svn_path,
                        'revision': parsed['revision'],
                        'author': parsed['author'],
                        'date': f"{parsed['date']} {parsed['time']}"
                    })
                    self.progress.emit(len(entries))
                
                if not entries:
                    raise Exception(self.tr('error_no_files'))
                
                self.log(self.tr('log_files_found', count=len(entries)))
                self.log(self.tr('log_creating_excel'))
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "SVN File Information"
                
                headers = ['File Name', 'Directory', 'Revision', 'Author', 'Commit Date']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                for row, entry in enumerate(entries, 2):
                    ws.cell(row=row, column=1, value=entry['file_name'])
                    ws.cell(row=row, column=2, value=entry['dir_path'])
                    ws.cell(row=row, column=3, value=entry['revision'])
                    ws.cell(row=row, column=4, value=entry['author'])
                    ws.cell(row=row, column=5, value=entry['date'])
                
                self.log(self.tr('log_saving_excel', path=self.excel_path))
                wb.save(self.excel_path)
                self.log(self.tr('log_excel_saved'))
                self.finished.emit(True, self.tr('log_export_success', count=len(entries)))
                
            except Exception as e:
                self.log(self.tr('log_svn_failed', error=str(e)))
                self.finished.emit(False, self.tr('error_svn', error=str(e)))
            
        except Exception as e:
            self.log(self.tr('log_program_failed', error=str(e)))
            self.finished.emit(False, self.tr('error_general', error=str(e))) 